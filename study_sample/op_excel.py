from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import PatternFill

EXSIT_CONSTRACT = [
    '广州百盈商务有限公司'
]

# 1.2.1 新建保存工作簿(覆盖创建）
def createWorkBook():
    wb=Workbook() # 新建工作簿
    wb.save('我的工作簿.xlsx') # 相对路径保存工作簿

# 1.2.2 读取保存工作簿
def loadWorkBook():
    wb=load_workbook('../doc/testdata/test文档.xlsx')
    wb.save('../doc/testdata/test文档.xlsx')

# 1.2.3 实例（批量建新工作簿）
def batchCreateWorkBook():
    for m in range(1,13):
        wb = Workbook()  # 新建工作簿
        wb.save('%d 月.xlsx'%m)# 保存工作簿

#loadWorkBook()
#batchCreateWorkBook()

def sqlGenerator(filePath):
    intersection = []
    notIn = []
    wb =load_workbook(filePath)
    # 获取当前活动的工作表
    # sheet = wb.active
    # print(sheet)
    # 获取所有工作表名
    sheetnames = wb.sheetnames
    #print(sheetnames)
    for sheetname in sheetnames:
        worksheet = wb[sheetname]
        print(worksheet)
        # 工作表['起始行号': '结束行号']或者工作表['起始行号: 结束行号']
        #s = worksheet['1:3']
        #print(s)
        for row in list(worksheet.rows)[1:]:
            l=[v.value for v in row]
            #print(row)
            #print(l[0])
            if l[0] in EXSIT_CONSTRACT:
                #print(l[0])
                intersection.append(l[0])
            else:
                # 设置单元格样式
                row[0].fill =  PatternFill("solid", fgColor='1874CD')
                notIn.append(l[0])
            print(row[0])

        print('--------交集--------')
        print(intersection)
        print('---------多的-----------')
        print(notIn)
        wb.save(filePath)

filePath = '../doc/testdata/test文档.xlsx'
sqlGenerator(filePath)

